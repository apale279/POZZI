"""Analizza formule Excel e colonne calcolabili/ottimizzabili in ECMO + ACC."""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

BASE = Path(r"c:\App_mie\ECMO")
FILES = {
    "ECMO": BASE / "DB VENOARTERIOSI CARDIO (1).xlsx",
    "ACC": BASE / "Nuovo DB ACC (2).xlsx",
}
OUT = BASE / "app" / "docs" / "calculable_fields_analysis.json"

# Colonne il cui nome suggerisce valore derivato (anche senza formula nel file)
DERIVED_NAME_PATTERNS = [
    (re.compile(r"^CALCOLO\s", re.I), "nome_esplicito_calcolo"),
    (re.compile(r"^BMI$", re.I), "bmi"),
    (re.compile(r"^ANNI$", re.I), "eta_anni"),
    (re.compile(r"^ANNO$", re.I), "anno_da_data"),
    (re.compile(r"^P/F$|^PF$", re.I), "rapporto_pf"),
    (re.compile(r"^S/F$", re.I), "rapporto_sf"),
    (re.compile(r"^SAPS", re.I), "score_saps"),
    (re.compile(r"^SOFA", re.I), "score_sofa"),
    (re.compile(r"^CPC", re.I), "score_cpc"),
    (re.compile(r"^GCS", re.I), "score_gcs"),
    (re.compile(r"^NPI", re.I), "score_npi"),
    (re.compile(r"%|PERCENT", re.I), "percentuale"),
    (re.compile(r"RATIO|RAPPORTO", re.I), "rapporto"),
    (re.compile(r"^CI$", re.I), "cardiac_index"),
    (re.compile(r"^SVRI?$", re.I), "indice_resistenze"),
    (re.compile(r"^DO2|^VO2", re.I), "derivato_ossigeno"),
]

DATE_COL_HINTS = re.compile(
    r"DATA|DATE|DN|NASCITA|ARRESTO|INGRESSO|START|END|DIMISSION",
    re.I,
)
WEIGHT_HEIGHT = {"PESO", "ALTEZZA", "HT", "HB", "BMI"}


def norm_header(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.startswith("Unnamed"):
        return None
    return s


def find_header_row(ws, max_scan=15) -> tuple[int, list[str]]:
    best_row, best_cols = 0, []
    for r in range(1, min(max_scan, ws.max_row or 1) + 1):
        vals = [norm_header(ws.cell(r, c).value) for c in range(1, (ws.max_column or 1) + 1)]
        cols = [v for v in vals if v]
        if len(cols) > len(best_cols):
            best_row, best_cols = r, cols
    return best_row, best_cols


def classify_column(name: str) -> list[str]:
    tags = []
    for rx, tag in DERIVED_NAME_PATTERNS:
        if rx.search(name):
            tags.append(tag)
    if DATE_COL_HINTS.search(name):
        tags.append("colonna_data")
    if name.upper() in WEIGHT_HEIGHT:
        tags.append("antropometria_lab")
    return tags


def extract_formula_refs(formula: str) -> list[str]:
    if not formula:
        return []
    return re.findall(r"[A-Z]{1,3}\d+", formula.upper())


def analyze_workbook(study: str, path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=True)
    result = {
        "file": path.name,
        "sheets": {},
        "formula_summary": [],
        "derived_columns": [],
        "optimization_hints": [],
    }

    formula_by_pattern: dict[str, int] = defaultdict(int)
    all_derived: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in ("PIVOT", "TENDINE SLIM"):
            continue
        ws = wb[sheet_name]
        header_row, headers = find_header_row(ws)
        if not headers:
            continue

        col_tags = {h: classify_column(h) for h in headers}
        sheet_formulas: list[dict] = []
        formula_cells = 0

        max_r = min((ws.max_row or header_row) + 50, header_row + 200)
        max_c = ws.max_column or len(headers)

        for r in range(header_row + 1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, Cell):
                    continue
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                formula_cells += 1
                f = val[1:].upper()
                # pattern semplificato
                pat = "other"
                if "YEAR(" in f or "ANNO(" in f:
                    pat = "year_from_date"
                elif "TODAY()" in f or "OGGI()" in f:
                    pat = "today"
                elif "DATEDIF" in f or "DATEDIF" in f or "YEARFRAC" in f:
                    pat = "age_or_interval"
                elif "/" in f and ("PESO" in f or "*" in f or "^" in f):
                    pat = "ratio_or_bmi"
                elif "IF(" in f:
                    pat = "conditional"
                elif "SUM(" in f or "AVERAGE(" in f or "MEDIA(" in f:
                    pat = "aggregate"
                elif "VLOOKUP" in f or "CERCA" in f:
                    pat = "lookup"
                formula_by_pattern[pat] += 1

                col_name = None
                if c <= len(headers):
                    col_name = headers[c - 1] if c - 1 < len(headers) else None
                sheet_formulas.append(
                    {
                        "cell": cell.coordinate,
                        "row": r,
                        "column": col_name,
                        "formula": val[:120],
                        "pattern": pat,
                    }
                )
                if len(sheet_formulas) >= 25:
                    break
            if len(sheet_formulas) >= 25:
                break

        derived_in_sheet = []
        for h, tags in col_tags.items():
            if tags and any(t not in ("colonna_data", "antropometria_lab") for t in tags):
                derived_in_sheet.append({"column": h, "tags": tags})
                all_derived.append({"sheet": sheet_name, "column": h, "tags": tags})

        # Ottimizzazione: ANNO ripetuto su molti fogli ECMO
        if study == "ECMO" and "ANNO" in headers:
            result["optimization_hints"].append(
                {
                    "sheet": sheet_name,
                    "hint": "colonna_ANNO_ripetuta",
                    "detail": "ANNO su ogni foglio ECMO — può derivare da data ingresso/arresto o anno studio",
                }
            )

        result["sheets"][sheet_name] = {
            "header_row": header_row,
            "column_count": len(headers),
            "headers_sample": headers[:20],
            "formula_cells_sampled": formula_cells,
            "formulas_sample": sheet_formulas[:15],
            "derived_columns": derived_in_sheet,
            "date_columns": [h for h in headers if "colonna_data" in col_tags.get(h, [])],
            "anthropometry": [h for h in headers if "antropometria_lab" in col_tags.get(h, [])],
        }

    result["formula_pattern_counts"] = dict(formula_by_pattern)
    result["derived_columns"] = all_derived

    # Regole cross-file
    result["recommended_rules"] = build_recommended_rules(study, result)
    wb.close()
    return result


def build_recommended_rules(study: str, data: dict) -> list[dict]:
    rules = []

    if study == "ECMO":
        rules.extend(
            [
                {
                    "id": "ecmo_bmi",
                    "targets": [{"sheet": "ANAGRAFICA", "column": "BMI"}],
                    "needs": ["PESO", "ALTEZZA"],
                    "formula": "peso_kg / (altezza_m)^2",
                    "priority": "high",
                },
                {
                    "id": "ecmo_age",
                    "targets": [{"sheet": "ANAGRAFICA", "column": "CALCOLO ETA"}, {"sheet": "ANAGRAFICA", "column": "ANNI"}],
                    "needs": ["DN"],
                    "formula": "eta da DN a data riferimento (ingresso ICU o arresto)",
                    "priority": "high",
                },
                {
                    "id": "ecmo_year_all_sheets",
                    "targets": "ANNO su quasi tutti i fogli (20×)",
                    "needs": ["DATA INGRESSO H", "DATA INGRESSO ICU", "START DATE", "DATA ARRESTO"],
                    "formula": "YEAR(data_riferimento_episodio)",
                    "priority": "high",
                    "optimization": "compilare 1× in anagrafica e propagare",
                },
                {
                    "id": "ecmo_pf_ratio",
                    "targets": [{"sheet": "24HRS ECLS ASSESSMENT", "column": "P/F"}],
                    "needs": ["PaO2/PO2", "FiO2"],
                    "formula": "PaO2 / (FiO2/100)",
                    "priority": "medium",
                },
                {
                    "id": "ecmo_identity_fanout",
                    "targets": "SDO, COGNOME, NOME, ELSO, RUN su ogni foglio",
                    "needs": ["core patient"],
                    "formula": "copia da anagrafica",
                    "priority": "high",
                    "optimization": "già in app (autoFilled)",
                },
            ]
        )
    else:
        rules.extend(
            [
                {
                    "id": "acc_bmi",
                    "targets": [{"sheet": "Anagrafica", "column": "BMI"}],
                    "needs": ["Peso", "Altezza"],
                    "formula": "peso / altezza²",
                    "priority": "high",
                },
                {
                    "id": "acc_year",
                    "targets": [{"sheet": "Anagrafica", "column": "Anno"}],
                    "needs": ["Data arresto", "Data ammissione"],
                    "formula": "YEAR(data arresto)",
                    "priority": "high",
                },
                {
                    "id": "acc_age",
                    "targets": [{"sheet": "Anagrafica", "column": "Età"}],
                    "needs": ["Data di nascita"],
                    "formula": "eta a data arresto",
                    "priority": "high",
                },
                {
                    "id": "acc_pf",
                    "targets": "P/F o equivalente su Ammissione, 6-12H, DAY 1-3",
                    "needs": ["PaO2", "FiO2 EGA"],
                    "formula": "PaO2/(FiO2/100)",
                    "priority": "medium",
                },
                {
                    "id": "acc_saps_sofa",
                    "targets": "SAPS-II, SOFA su timepoint",
                    "needs": "variabili componenti (se presenti come sub-score)",
                    "formula": "score composito — verificare se calcolati in Excel",
                    "priority": "low",
                },
            ]
        )
    return rules


def main():
    out = {}
    for study, path in FILES.items():
        print(f"Analyzing {study}...")
        out[study] = analyze_workbook(study, path)

    # Cross-study
    out["cross_study"] = {
        "overlap_calculable": [
            "patient.sdo/cognome/nome — fan-out (non calcolo ma ottimizzazione)",
            "blood_gas.pf_ratio — PaO2 + FiO2",
            "study.year (ANNO) — da date episodio",
            "patient.bmi — peso + altezza",
            "patient.age — DN + data riferimento",
        ],
        "ecmo_specific": [
            "RUN ripetuto — propagare da selezione run",
            "CALCOLO ETA / ANNI solo ANAGRAFICA ECMO",
        ],
        "acc_specific": [
            "Timepoint DAY 1-3 — stesse colonne EGA/vent ripetute (estrazione 1 template)",
            "CPC outcome — derivato da esame neurologico non automatico",
        ],
    }

    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Written {OUT}")

    # Print summary
    for study in ("ECMO", "ACC"):
        d = out[study]
        print(f"\n=== {study} ===")
        print(f"Formula patterns: {d.get('formula_pattern_counts')}")
        print(f"Derived columns (by name): {len(d.get('derived_columns', []))}")
        for r in d.get("recommended_rules", []):
            print(f"  - [{r['priority']}] {r['id']}: {r.get('formula')}")


if __name__ == "__main__":
    main()
